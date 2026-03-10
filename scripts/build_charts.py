"""
Build Excel charts + PNG images from urea trade data.
Reads africa_urea_trade_latest.csv and produces:
  - africa_urea_analysis_YYYYMMDD.xlsx (Excel with charts)
  - charts/top_importers.png
  - charts/import_trends.png
  - charts/net_trade_position.png
  - charts/heatmap_value.png
  - charts/heatmap_weight.png

Usage:
    python scripts/build_charts.py
"""

import sys
import subprocess
from pathlib import Path
from datetime import datetime, timezone

for pkg in ['pandas', 'openpyxl', 'plotly', 'kaleido']:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent.parent / "data"
CHART_DIR = DATA_DIR / "charts"
CHART_DIR.mkdir(parents=True, exist_ok=True)

AFC_BLUE = '#1F4E79'
AFC_LIGHT = '#2E75B6'
AFC_ORANGE = '#ED7D31'
AFC_GREEN = '#70AD47'
AFC_GREY = '#A5A5A5'

df = pd.read_csv(DATA_DIR / "africa_urea_trade_latest.csv")
df['trade_value_1000usd'] = pd.to_numeric(df['trade_value_1000usd'], errors='coerce')
df['quantity_kg'] = pd.to_numeric(df['quantity_kg'], errors='coerce')
df['year'] = df['year'].astype(int)

imports = df[df['trade_flow'] == 'Import'].copy()
exports = df[df['trade_flow'] == 'Export'].copy()
years = sorted(imports['year'].unique())
latest_year = imports['year'].max()

print(f"Building charts from {len(df)} records...")
print(f"Years: {years[0]}-{years[-1]}, Latest: {latest_year}")

# ============================================================
# PLOTLY CHART 1: Top 10 Importers
# ============================================================
latest_imp = imports[imports['year'] == latest_year].sort_values('trade_value_1000usd', ascending=True).tail(10)

fig1 = go.Figure(go.Bar(
    x=latest_imp['trade_value_1000usd'],
    y=latest_imp['country_name'],
    orientation='h',
    marker_color=AFC_BLUE,
    text=latest_imp['trade_value_1000usd'].apply(lambda x: f'${x:,.0f}K'),
    textposition='outside'
))
fig1.update_layout(
    title=f'Top 10 African Urea Importers ({latest_year})',
    xaxis_title='Trade Value ($K USD)',
    yaxis_title='',
    template='plotly_white',
    font=dict(family='Arial', size=12),
    height=500, width=900,
    margin=dict(l=150, r=100)
)
fig1.write_image(str(CHART_DIR / "top_importers.png"), scale=2)
print("  Saved: charts/top_importers.png")

# ============================================================
# PLOTLY CHART 2: Import Trends (top 5)
# ============================================================
total_by_country = imports.groupby('country_name')['trade_value_1000usd'].sum().sort_values(ascending=False)
top5 = total_by_country.head(5).index.tolist()
trend_data = imports[imports['country_name'].isin(top5)].copy()

fig2 = px.line(
    trend_data, x='year', y='trade_value_1000usd', color='country_name',
    labels={'trade_value_1000usd': 'Trade Value ($K USD)', 'year': 'Year', 'country_name': 'Country'},
    title='Urea Import Trends — Top 5 African Importers (2010-2024)',
    template='plotly_white',
    color_discrete_sequence=[AFC_BLUE, AFC_LIGHT, AFC_ORANGE, AFC_GREEN, AFC_GREY]
)
fig2.update_layout(
    font=dict(family='Arial', size=12),
    height=500, width=900,
    legend=dict(title='', orientation='h', yanchor='bottom', y=-0.25, xanchor='center', x=0.5)
)
fig2.update_traces(line=dict(width=2.5))
fig2.write_image(str(CHART_DIR / "import_trends.png"), scale=2)
print("  Saved: charts/import_trends.png")

# ============================================================
# PLOTLY CHART 3: Net Trade Position
# ============================================================
imp_totals = imports.groupby('country_name')['trade_value_1000usd'].sum()
exp_totals = exports.groupby('country_name')['trade_value_1000usd'].sum()
all_countries = sorted(set(imp_totals.index) | set(exp_totals.index))

pos_data = []
for c in all_countries:
    imp_val = imp_totals.get(c, 0)
    exp_val = exp_totals.get(c, 0)
    pos_data.append({'country': c, 'imports': imp_val, 'exports': exp_val, 'net': imp_val - exp_val})

pos_df = pd.DataFrame(pos_data).sort_values('net')
top_pos = pd.concat([pos_df.head(10), pos_df.tail(10)]).drop_duplicates()

colors = [AFC_GREEN if x < 0 else AFC_ORANGE for x in top_pos['net']]

fig3 = go.Figure(go.Bar(
    x=top_pos['net'],
    y=top_pos['country'],
    orientation='h',
    marker_color=colors,
    text=top_pos['net'].apply(lambda x: f'${x:,.0f}K'),
    textposition='outside'
))
fig3.update_layout(
    title='Net Urea Trade Position — African Countries (Cumulative 2010-2024)',
    xaxis_title='Net Long/Short ($K USD) — Green = Short (Producer), Orange = Long (Importer)',
    yaxis_title='',
    template='plotly_white',
    font=dict(family='Arial', size=11),
    height=600, width=1000,
    margin=dict(l=150, r=120)
)
fig3.add_vline(x=0, line_dash="dash", line_color="grey")
fig3.write_image(str(CHART_DIR / "net_trade_position.png"), scale=2)
print("  Saved: charts/net_trade_position.png")

# ============================================================
# PLOTLY CHART 4: Heatmap by Value
# ============================================================
imp_pivot_val = imports.pivot_table(index='country_name', columns='year', values='trade_value_1000usd', aggfunc='sum')
imp_pivot_val = imp_pivot_val.loc[imp_pivot_val.sum(axis=1).sort_values(ascending=False).index]

fig4 = go.Figure(go.Heatmap(
    z=imp_pivot_val.values,
    x=[str(y) for y in imp_pivot_val.columns],
    y=imp_pivot_val.index,
    colorscale='Greens',
    text=imp_pivot_val.applymap(lambda x: f'{x:,.0f}' if pd.notna(x) else '').values,
    texttemplate='%{text}',
    textfont=dict(size=7),
    hoverongaps=False,
    colorbar=dict(title='$K USD')
))
fig4.update_layout(
    title='African Urea Imports — Trade Value ($K USD)',
    template='plotly_white',
    font=dict(family='Arial', size=10),
    height=max(600, len(imp_pivot_val) * 16),
    width=900,
    yaxis=dict(autorange='reversed'),
    margin=dict(l=160)
)
fig4.write_image(str(CHART_DIR / "heatmap_value.png"), scale=2)
print("  Saved: charts/heatmap_value.png")

# ============================================================
# PLOTLY CHART 5: Heatmap by Weight
# ============================================================
imports_wt = imports.copy()
imports_wt['quantity_mt'] = imports_wt['quantity_kg'] / 1000

imp_pivot_wt = imports_wt.pivot_table(index='country_name', columns='year', values='quantity_mt', aggfunc='sum')
imp_pivot_wt = imp_pivot_wt.loc[imp_pivot_wt.sum(axis=1).sort_values(ascending=False).index]

fig5 = go.Figure(go.Heatmap(
    z=imp_pivot_wt.values,
    x=[str(y) for y in imp_pivot_wt.columns],
    y=imp_pivot_wt.index,
    colorscale='Blues',
    text=imp_pivot_wt.applymap(lambda x: f'{x:,.0f}' if pd.notna(x) else '').values,
    texttemplate='%{text}',
    textfont=dict(size=7),
    hoverongaps=False,
    colorbar=dict(title='MT')
))
fig5.update_layout(
    title='African Urea Imports — Quantity (Metric Tonnes)',
    template='plotly_white',
    font=dict(family='Arial', size=10),
    height=max(600, len(imp_pivot_wt) * 16),
    width=900,
    yaxis=dict(autorange='reversed'),
    margin=dict(l=160)
)
fig5.write_image(str(CHART_DIR / "heatmap_weight.png"), scale=2)
print("  Saved: charts/heatmap_weight.png")

# ============================================================
# EXCEL WORKBOOK (same as before)
# ============================================================
print("\nBuilding Excel workbook...")

wb = Workbook()
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=10)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(italic=True, size=10, color='808080')
ALT_ROW = PatternFill('solid', fgColor='D6E4F0')

def style_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Sheet 1: Top Importers
ws1 = wb.active
ws1.title = "Top Importers"
ws1.sheet_properties.tabColor = "1F4E79"
latest_imp_xl = imports[imports['year'] == latest_year].sort_values('trade_value_1000usd', ascending=False).head(10)
ws1['A1'] = f'Top 10 African Urea Importers ({latest_year})'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:E1')
style_header(ws1, 3, ['Country', 'Value ($K USD)', 'Quantity (MT)', 'Implied Price ($/MT)', 'Data Note'])
for idx, (_, row) in enumerate(latest_imp_xl.iterrows(), 4):
    ws1.cell(row=idx, column=1, value=row['country_name'])
    ws1.cell(row=idx, column=2, value=round(row['trade_value_1000usd'], 1)).number_format = '#,##0.0'
    qty_mt = round(row['quantity_kg'] / 1000, 0) if pd.notna(row['quantity_kg']) and row['quantity_kg'] > 0 else None
    ws1.cell(row=idx, column=3, value=qty_mt).number_format = '#,##0'
    if qty_mt and qty_mt > 0:
        ws1.cell(row=idx, column=4, value=f'=B{idx}*1000/C{idx}').number_format = '$#,##0'
    else:
        ws1.cell(row=idx, column=4, value='N/A')
        ws1.cell(row=idx, column=5, value='Qty not reported')
        ws1.cell(row=idx, column=5).font = Font(italic=True, color='808080', size=9)
    if idx % 2 == 0:
        for c in range(1, 6):
            ws1.cell(row=idx, column=c).fill = ALT_ROW
for col, w in [(1, 22), (2, 18), (3, 16), (4, 20), (5, 18)]:
    ws1.column_dimensions[get_column_letter(col)].width = w

xl_chart1 = BarChart()
xl_chart1.type = "bar"
xl_chart1.style = 10
xl_chart1.title = f"Top 10 African Urea Importers ({latest_year})"
xl_chart1.y_axis.title = "Country"
xl_chart1.x_axis.title = "Trade Value ($K USD)"
xl_chart1.height = 15
xl_chart1.width = 25
d1 = Reference(ws1, min_col=2, min_row=3, max_row=3 + len(latest_imp_xl))
c1 = Reference(ws1, min_col=1, min_row=4, max_row=3 + len(latest_imp_xl))
xl_chart1.add_data(d1, titles_from_data=True)
xl_chart1.set_categories(c1)
xl_chart1.series[0].graphicalProperties.solidFill = "1F4E79"
xl_chart1.legend = None
ws1.add_chart(xl_chart1, "G3")

# Sheet 2: Import Trends
ws2 = wb.create_sheet("Import Trends")
ws2.sheet_properties.tabColor = "2E75B6"
ws2['A1'] = 'Urea Import Trends — Top 5 African Importers (2010-2024)'
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A1:G1')
ws2.cell(row=3, column=1, value='Year').font = HDR_FONT
ws2.cell(row=3, column=1).fill = HDR_FILL
for j, country in enumerate(top5, 2):
    cell = ws2.cell(row=3, column=j, value=country)
    cell.font = Font(bold=True, color='FFFFFF', size=9)
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center')
for i, year in enumerate(years, 4):
    ws2.cell(row=i, column=1, value=year)
    for j, country in enumerate(top5, 2):
        val = imports[(imports['country_name'] == country) & (imports['year'] == year)]['trade_value_1000usd']
        ws2.cell(row=i, column=j, value=round(val.values[0], 1) if len(val) > 0 and pd.notna(val.values[0]) else None)
        ws2.cell(row=i, column=j).number_format = '#,##0.0'
ws2.column_dimensions['A'].width = 8
for c in range(2, len(top5) + 2):
    ws2.column_dimensions[get_column_letter(c)].width = 16

xl_chart2 = LineChart()
xl_chart2.style = 10
xl_chart2.title = "Urea Import Value ($K USD) — Top 5 African Importers"
xl_chart2.y_axis.title = "Trade Value ($K USD)"
xl_chart2.x_axis.title = "Year"
xl_chart2.y_axis.numFmt = '#,##0'
xl_chart2.height = 18
xl_chart2.width = 28
d2 = Reference(ws2, min_col=2, max_col=len(top5) + 1, min_row=3, max_row=3 + len(years))
c2 = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(years))
xl_chart2.add_data(d2, titles_from_data=True)
xl_chart2.set_categories(c2)
xl_colors = ['1F4E79', '2E75B6', 'ED7D31', '70AD47', 'A5A5A5']
for idx, s in enumerate(xl_chart2.series):
    s.graphicalProperties.line.width = 25000
    s.graphicalProperties.line.solidFill = xl_colors[idx % len(xl_colors)]
ws2.add_chart(xl_chart2, "A" + str(4 + len(years) + 2))

# Sheet 3: Net Urea Trade Position
ws3 = wb.create_sheet("Net Urea Trade Position")
ws3.sheet_properties.tabColor = "ED7D31"
ws3['A1'] = 'Net Urea Trade Position — African Countries (Cumulative 2010-2024)'
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A1:F1')
ws3['A2'] = 'Long = net importer (buys urea). Short = net exporter (produces/sells urea).'
ws3['A2'].font = SUBTITLE_FONT
ws3.merge_cells('A2:F2')
style_header(ws3, 4, ['Country', 'Total Imports ($K)', 'Total Exports ($K)', 'Net Long/Short ($K)', 'Position', 'Producer Flag'])

country_data = []
for c in all_countries:
    iv = imp_totals.get(c, 0)
    ev = exp_totals.get(c, 0)
    country_data.append((c, iv, ev, iv - ev))
country_data.sort(key=lambda x: x[3])

SHORT_FILL = PatternFill('solid', fgColor='C6EFCE')
MAYBE_FILL = PatternFill('solid', fgColor='FFEB9C')

for idx, (country, iv, ev, net) in enumerate(country_data, 5):
    ws3.cell(row=idx, column=1, value=country)
    ws3.cell(row=idx, column=2, value=round(iv, 1)).number_format = '#,##0.0'
    ws3.cell(row=idx, column=3, value=round(ev, 1)).number_format = '#,##0.0'
    ws3.cell(row=idx, column=4, value=f'=B{idx}-C{idx}').number_format = '#,##0.0'
    if net < 0:
        ws3.cell(row=idx, column=5, value='SHORT')
        ws3.cell(row=idx, column=5).font = Font(bold=True, color='006100')
        ws3.cell(row=idx, column=5).fill = SHORT_FILL
    else:
        ws3.cell(row=idx, column=5, value='LONG')
        ws3.cell(row=idx, column=5).font = Font(color='9C0006')
    ws3.cell(row=idx, column=5).alignment = Alignment(horizontal='center')
    if ev > 100000:
        ws3.cell(row=idx, column=6, value='Major Producer')
        ws3.cell(row=idx, column=6).font = Font(bold=True, color='006100')
        ws3.cell(row=idx, column=6).fill = SHORT_FILL
    elif ev > 10000:
        ws3.cell(row=idx, column=6, value='Significant Exporter')
        ws3.cell(row=idx, column=6).font = Font(bold=True, color='006100')
        ws3.cell(row=idx, column=6).fill = SHORT_FILL
    elif ev > 1000:
        ws3.cell(row=idx, column=6, value='Possible Producer / Re-exporter')
        ws3.cell(row=idx, column=6).font = Font(color='9C6500')
        ws3.cell(row=idx, column=6).fill = MAYBE_FILL
    elif ev > 100:
        ws3.cell(row=idx, column=6, value='Minor Exporter')
        ws3.cell(row=idx, column=6).font = Font(color='9C6500')
        ws3.cell(row=idx, column=6).fill = MAYBE_FILL
    else:
        ws3.cell(row=idx, column=6, value='Pure Importer')

for col, w in [(1, 22), (2, 20), (3, 20), (4, 22), (5, 12), (6, 28)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

# Sheets 4-6: Heatmaps + Raw Data (values written directly, no Excel charts needed since Plotly covers it)
ws4 = wb.create_sheet("Heatmap (Value)")
ws4.sheet_properties.tabColor = "70AD47"
ws4['A1'] = 'See charts/heatmap_value.png for visual. Raw data below.'
ws4['A1'].font = SUBTITLE_FONT
ws4.cell(row=3, column=1, value='Country').font = HDR_FONT
ws4.cell(row=3, column=1).fill = HDR_FILL
imp_countries_val = imports.groupby('country_name')['trade_value_1000usd'].sum().sort_values(ascending=False).index.tolist()
for j, year in enumerate(years, 2):
    cell = ws4.cell(row=3, column=j, value=year)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
for i, country in enumerate(imp_countries_val, 4):
    ws4.cell(row=i, column=1, value=country)
    for j, year in enumerate(years, 2):
        val_s = imports[(imports['country_name'] == country) & (imports['year'] == year)]['trade_value_1000usd']
        val = val_s.values[0] if len(val_s) > 0 and pd.notna(val_s.values[0]) else None
        ws4.cell(row=i, column=j, value=round(val, 0) if val else None).number_format = '#,##0'
ws4.column_dimensions['A'].width = 22

ws5 = wb.create_sheet("Heatmap (Weight)")
ws5.sheet_properties.tabColor = "2E75B6"
ws5['A1'] = 'See charts/heatmap_weight.png for visual. Raw data in MT below.'
ws5['A1'].font = SUBTITLE_FONT
imports_wt = imports.copy()
imports_wt['quantity_mt'] = imports_wt['quantity_kg'] / 1000
wt_countries = imports_wt.groupby('country_name')['quantity_mt'].sum().sort_values(ascending=False).index.tolist()
ws5.cell(row=3, column=1, value='Country').font = HDR_FONT
ws5.cell(row=3, column=1).fill = HDR_FILL
for j, year in enumerate(years, 2):
    cell = ws5.cell(row=3, column=j, value=year)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
for i, country in enumerate(wt_countries, 4):
    ws5.cell(row=i, column=1, value=country)
    for j, year in enumerate(years, 2):
        val_s = imports_wt[(imports_wt['country_name'] == country) & (imports_wt['year'] == year)]['quantity_mt']
        val = val_s.values[0] if len(val_s) > 0 and pd.notna(val_s.values[0]) and val_s.values[0] > 0 else None
        ws5.cell(row=i, column=j, value=round(val, 0) if val else None).number_format = '#,##0'
ws5.column_dimensions['A'].width = 22

ws6 = wb.create_sheet("Raw Data")
ws6.sheet_properties.tabColor = "A5A5A5"
for j, col in enumerate(df.columns, 1):
    cell = ws6.cell(row=1, column=j, value=col)
    cell.font = HDR_FONT
    cell.fill = PatternFill('solid', fgColor='404040')
for i, (_, row) in enumerate(df.iterrows(), 2):
    for j, col in enumerate(df.columns, 1):
        ws6.cell(row=i, column=j, value=row[col])

timestamp = datetime.now(timezone.utc).strftime("%Y%m%d")
outpath = DATA_DIR / f"africa_urea_analysis_{timestamp}.xlsx"
wb.save(outpath)
import shutil
shutil.copy2(outpath, DATA_DIR / "africa_urea_analysis_latest.xlsx")
print(f"\n  Saved Excel: {outpath}")
print(f"  Sheets: {wb.sheetnames}")
print("\nDone!")
